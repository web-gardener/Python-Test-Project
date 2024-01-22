# app.py
from flask import Flask, request, jsonify, abort
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import openpyxl

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///books.db'
db = SQLAlchemy(app)


class Author(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    birth_date = db.Column(db.String(10), nullable=False)

class Book(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    barcode = db.Column(db.String(255))
    title = db.Column(db.String(255), nullable=False)
    publish_year = db.Column(db.Integer, nullable=False)
    author_id = db.Column(db.Integer, db.ForeignKey('author.id'), nullable=False)
    author = db.relationship('Author', backref=db.backref('books', lazy=True))

class Storing(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    book_id = db.Column(db.Integer, db.ForeignKey('book.id'), nullable=False)
    book = db.relationship('Book', backref=db.backref('storing_info', lazy=True))
    quantity = db.Column(db.Integer, nullable=False)
    date = db.Column(db.DateTime, default=db.func.now())


# Ping endpoint
@app.route('/ping')
def ping():
    return jsonify({"message": "pong"}), 200

# Authors endpoints
@app.route('/author', methods=['POST'])
def add_author():
    data = request.json

    if not all(key in data for key in ('name', 'birth_date')):
        return jsonify({"message": "Wrong request structure"}), 422

    new_author = Author(**data)
    db.session.add(new_author)
    db.session.commit()
    return jsonify({"key": new_author.id}), 201

@app.route('/author/<int:key>', methods=['GET'])
def get_author(key):
    try: 
        author = Author.query.get(key)

        return jsonify({"key": author.id, "name": author.name}), 200
    except:
        return jsonify({"message": "Entity not found"}), 404

# Books endpoints
@app.route('/book', methods=['POST'])
def add_book():
    data = request.json

    if not all(key in data for key in ('barcode', 'title', 'publish_year', 'author_id')):
        return jsonify({"message": "Wrong request structure"}), 422

    new_book = Book(**data)
    db.session.add(new_book)
    db.session.commit()
    return jsonify({"key": new_book.id}), 201

@app.route('/book/<int:key>', methods=['GET'])
def get_book(key):
    try:
        book = Book.query.get(key)

        result = {
            "key": book.id,
            "barcode": book.barcode,
            "title": book.title,
            "publish_year": book.publish_year,
            "author": {"name": book.author.name, "birth_date": book.author.birth_date},
            "quantity": book.storing_info[-1].quantity if book.storing_info else 0
        }
        return jsonify(result), 200
    except:
        return jsonify({"message": "Entity not found"}), 404


@app.route('/book', methods=['GET'])
def search_book():
    barcode = request.args.get('barcode')
    books = Book.query.filter_by(barcode=barcode).order_by(Book.barcode).all()

    result = {
        "found": len(books),
        "items": [
            {
                "key": book.id,
                "barcode": book.barcode,
                "title": book.title,
                "publish_year": book.publish_year,
                "author": {"name": book.author.name, "birth_date": book.author.birth_date},
                "quantity": book.storing_info[-1].quantity if book.storing_info else 0
            } for book in books
        ]
    }
    return jsonify(result), 200

# Storing endpoints
@app.route('/leftover/add', methods=['POST'])
def add_leftover():
    data = request.json

    if not all(key in data for key in ('barcode', 'quantity')):
        return jsonify({"message": "Wrong request structure"}), 422

    barcode = data['barcode']
    quantity = data['quantity']

    book = Book.query.filter_by(barcode=barcode).first()
    if not book:
        return jsonify({"message": f"Book with barcode {barcode} not found"}), 404

    new_leftover = Storing(book=book, quantity=quantity, date=datetime.utcnow())
    db.session.add(new_leftover)
    db.session.commit()
    return jsonify({"key": new_leftover.id}), 201
@app.route('/leftover/remove', methods=['POST'])
def remove_leftover():
    data = request.json

    if not all(key in data for key in ('barcode', 'quantity')):
        return jsonify({"message": "Wrong request structure"}), 422

    barcode = data['barcode']
    quantity = -data['quantity']  # Quantity for removal is negated

    book = Book.query.filter_by(barcode=barcode).first()
    if not book:
        return jsonify({"message": f"Book with barcode {barcode} not found"}), 404

    new_leftover = Storing(book=book, quantity=quantity, date=datetime.utcnow())
    db.session.add(new_leftover)
    db.session.commit()
    return jsonify({"key": new_leftover.id}), 201

@app.route('/leftovers/bulk', methods=['POST'])
def bulk_leftovers():
    try:
        file = request.files['file']
        upload_data = []

        if file.filename.endswith('.xlsx') or file.filename.endswith('.txt'):
            if file.filename.endswith('.xlsx'):
                data = openpyxl.load_workbook(file).active

                for row in range(1, data.max_row + 1):
                    barcode = data.cell(row=row, column=1).value
                    quantity = data.cell(row=row, column=2).value
                
                    if not barcode:
                        continue

                    book = Book.query.filter_by(barcode=barcode).first()

                    if not book:
                        return jsonify({"message": f"Book with barcode {barcode} not found in row {row}"}), 404

                    try:
                        quantity = int(quantity)
                    except ValueError:
                        return jsonify({"message": f"Invalid quantity format in row {row}"}), 400
                    
                    upload_data.append({barcode: barcode, quantity: quantity})
                return jsonify({"message": "File Uploaded Successfully"}), 200
            else:
                text_content = file.read().decode('utf-8')
                lines = text_content.splitlines()

                for i, line in enumerate(lines):
                    if len(line) < 3:
                        return jsonify({"message": f"Invalid line format in row {i + 1}"}), 400

                    line_type = line[:3]
                    line_data = line[3:].strip()

                    if line_type == "BRC":
                        if not line_data:
                            continue  # Skip rows with empty barcode
                        book = Book.query.filter_by(barcode=line_data).first()

                        if not book:
                            return jsonify({"message": f"Book with barcode {line_data} not found in row {i + 1}"}), 404

                    elif line_type == "QNT":
                        try:
                            quantity = int(line_data)
                        except ValueError:
                            return jsonify({"message": f"Invalid quantity format in row {i + 1}"}), 400

                        upload_data.append({"barcode": line_data, "quantity": quantity})

            for upload_data_item in upload_data:
                barcode = upload_data_item['barcode']
                quantity = upload_data_item['quantity']
                book = Book.query.filter_by(barcode=barcode).first()
                new_leftover = Storing(book=book, quantity=quantity, date=datetime.utcnow())
                db.session.add(new_leftover)
                db.session.commit()
            
            return jsonify({"message": f"Data successfully uploaded"}), 200

        else:
            return jsonify({"message": "Invalid file format. Please upload a XLSX or TXT file."}), 400

    except Exception as e:
        error_message = {"message": f"Error processing file: {str(e)}"}
        return jsonify(error_message), 400

@app.route('/history', methods=['GET'])
def get_history():
    if not request.args.get('start') or not request.args.get('end') or not request.args.get('book'):
        return jsonify({"message": "Wrong request structure"}), 422

    start_date_str = request.args.get('start')
    end_date_str = request.args.get('end')
    book_key = request.args.get('book')

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

    filters = []
    if start_date:
        filters.append(Storing.date >= start_date)
    if end_date:
        filters.append(Storing.date <= end_date)
    if book_key:
        filters.append(Storing.book_id == book_key)

    history_data = []
    book = Book.query.filter_by(id=book_key).first()
    sorted_storing_data = Storing.query.order_by(Storing.date.desc())
        
    history_entries = []
    start_balance = 0
    end_balance = 0

    for storing in sorted_storing_data:
        
        if not start_date or storing.date >= start_date:
            start_balance += storing.quantity

        if not end_date or storing.date <= end_date:
            end_balance += storing.quantity

        if storing.date >= start_date and storing.date <= end_date:
            history_entries.append({
                "date": storing.date.strftime('%Y-%m-%d %H:%M:%S'),
                "quantity": storing.quantity
            })

    history_data.append({
        "book": {"key": book.id, "title": book.title, "barcode": book.barcode},
        "start_balance": start_balance,
        "end_balance": end_balance,
        "history": history_entries
    })

    return jsonify(history_data), 200


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=False)
